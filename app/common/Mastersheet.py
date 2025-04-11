from io import BytesIO
from zipfile import ZipFile
from typing import BinaryIO
from openpyxl.worksheet.worksheet import Worksheet
from openpyxl import Workbook, load_workbook


class Mastersheet:
    def __init__(
        self, mastersheet_cols: dict, macro_file_path: str, mastersheet: BinaryIO | str
    ):
        self.MASTER_SHEET_COLS = mastersheet_cols

        self.macro_wb = load_workbook(filename=macro_file_path, keep_vba=True)
        self.macro_ws = self.macro_wb.active

        self.mastersheet_ws = load_workbook(filename=mastersheet).worksheets[0]
        self._parse_mastersheet(self.mastersheet_ws)
        self.mastersheet_dict = self._create_mastersheet_dict(self.mastersheet_ws)
        self.mastersheet_header = [cell.value for cell in self.mastersheet_ws[1]]

        self.zip_buffer = BytesIO()

    def _parse_mastersheet(self, mastersheet: Worksheet):
        """Parses the master excel sheet to find header columns"""
        for cell in mastersheet[1]:
            col_name = cell.value.lower() if cell.value else None

            if col_name in self.MASTER_SHEET_COLS:
                self.MASTER_SHEET_COLS[col_name] = cell.col_idx
        return

    def _create_mastersheet_dict(
        self, mastersheet: Worksheet, key="upc", shouldBreak=True
    ):
        """Creates a dictionary out of mastersheet with a unique key and row data as values"""
        mastersheet_dict = {}
        for row in mastersheet.iter_rows(min_row=2, min_col=1):
            if row[self.MASTER_SHEET_COLS[key] - 1].value is None:
                if shouldBreak:
                    break
                else:
                    continue

            mastersheet_dict[int(row[self.MASTER_SHEET_COLS[key] - 1].value)] = [
                cell.value for cell in row
            ]

        return mastersheet_dict

    def get_zip_buffer(self):
        self.zip_buffer.seek(0)
        return self.zip_buffer

    def _add_wb_to_zip(self, wb: Workbook, filename: str):
        with ZipFile(self.zip_buffer, "a") as zip_file:
            with BytesIO() as wb_io:
                wb.save(wb_io)
                zip_file.writestr(filename, wb_io.getvalue())

    def save_zip(self, zip_file_path: str):
        self.zip_buffer.seek(0)
        with open(zip_file_path, "wb") as f:
            f.write(self.zip_buffer.read())

    def save_wb(self, filename: str):
        self.macro_wb.save(filename)
        return
