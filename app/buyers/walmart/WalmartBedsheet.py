from io import BytesIO
from fastapi import UploadFile
from typing import BinaryIO, List
from zipfile import ZipFile

from openpyxl import Workbook, load_workbook
from openpyxl.worksheet.worksheet import Worksheet
from openpyxl.styles import Font
from app.utils.format import (
    apply_borders_and_alignment,
    format_number,
    adjust_column_widths,
    format_header,
    get_date,
)
from app.utils.helpers import get_po_sheet


class WalmartBedsheet:
    """
    Walmart Bedsheet class to handle the creation of mastersheet and macrofile for Walmart.
    """

    MACRO_FILE_PATH = "macro/walmart-bedsheet.xlsm"
    MASTERSHEET_WS_NAME = "300TC & 400TC Format"
    MASTER_SHEET_COLS = {
        "key": None,
        "dc location": None,
        "material number": None,
        "qty": None,
        "po number": None,
        "remarks": None,
        "upc": None,
    }

    PO_SHEET_INDEX = 3
    PO_SHEET_COLS = {
        "key": None,
        "po no": None,
        "remarks": None,
        "ordering qty": None,
        "supplier ship date": None,
        "supplier cancel date": None,
        "quality": None,
        "dc location": None,
        "prime item nbr": None,
    }

    def __init__(self, mastersheet: BinaryIO, po_file: BinaryIO):
        self.last_po: any = ""

        self.mastersheet_ws = load_workbook(filename=mastersheet, data_only=True)[
            self.MASTERSHEET_WS_NAME
        ]
        self.mastersheet_headers = [cell.value for cell in self.mastersheet_ws[1]]
        self.mastersheet_dict = self._parse_mastersheet(self.mastersheet_ws)

        # the macro wb to populate and return to user
        self.macro_wb = load_workbook(filename=self.MACRO_FILE_PATH, keep_vba=True)
        self.macro_ws = self.macro_wb.active

        # A new mastersheet wb that we need to return to the user.
        self.wb = Workbook()
        self.ws = self.wb.active
        self.ws.append(self.mastersheet_headers)

        # Calamine sheet
        self.po_sheet = get_po_sheet(po_file, self.PO_SHEET_INDEX)
        self._parse_po(self.po_sheet)
        self.po_sheet = self._sort_po(self.po_sheet)

    def _sort_po(self, ws: List[List]):
        sorted_list = sorted(
            ws[2:],
            key=lambda x: (
                float(x[self.PO_SHEET_COLS["po no"]])
                if x[self.PO_SHEET_COLS["po no"]]
                else float("inf")
            ),
        )
        sorted_list_headers = ws[:2] + sorted_list
        return sorted_list_headers

    def process(self):
        """Main method to process excel files"""

        """ 1. Loop through QTY col in PO
            2. If QTY is non-zero, get the mastersheet record from the dict using the key
            3. Now you have the PO row and mastersheet row.
            4. Get respective data items required for macro from both these rows.
            5. Push the record in new mastersheet wb with QTY, PO NO, and Remarks
            5. Create a macro-wb-record and push the record in the macrosheet.
            6. After loop is finished, return the zip buffer with wb and macro_wb files.
        """
        # Start with the first record (first row is blank, second row for headers)
        for row in self.po_sheet[2:]:
            qty = row[self.PO_SHEET_COLS["ordering qty"]]
            if qty == "" or qty is None:
                break

            if qty == 0:
                continue

            key = row[self.PO_SHEET_COLS["key"]]
            po_number = row[self.PO_SHEET_COLS["po no"]]
            self.current_po = po_number
            remarks = row[self.PO_SHEET_COLS["remarks"]]
            supplier_ship_date = row[self.PO_SHEET_COLS["supplier ship date"]]
            supplier_cancel_date = row[self.PO_SHEET_COLS["supplier cancel date"]]
            quality = row[self.PO_SHEET_COLS["quality"]]
            pis: int
            if quality == "400 TC":
                pis = 10000016237
            else:
                pis = 10000016236
            ordering_qty = row[self.PO_SHEET_COLS["ordering qty"]]
            dc_location = row[self.PO_SHEET_COLS["dc location"]]
            item_number = row[self.PO_SHEET_COLS["prime item nbr"]]

            # get the mastersheet record from the dict using corresponding key
            ms_row = self.mastersheet_dict.get(
                key, ["#N/A"] * len(self.mastersheet_headers)
            )

            ms_row[self.MASTER_SHEET_COLS["qty"] - 1] = ordering_qty
            ms_row[self.MASTER_SHEET_COLS["po number"] - 1] = po_number
            ms_row[self.MASTER_SHEET_COLS["remarks"] - 1] = remarks

            mat_code = ms_row[self.MASTER_SHEET_COLS["material number"] - 1]

            self.ws.append(ms_row)

            # add record to macrows
            if self.current_po != self.last_po:
                # Add empty row
                self.macro_ws.append([""])
                self.last_po = self.current_po

            macro_row = [
                po_number,  # po number
                "",  # so number
                100003,  # sold to party
                100003,  # ship to party
                1014,  # payment term
                "",  # inco term
                "MUNDRA",  # inco term 2
                100003,  # end customer
                "YES", # E2E clarity
                "RETAIL",  # channel type
                "NIL",  # sub channel type
                "REPLENISHMENT",  # order type
                supplier_ship_date,  # ship start date
                supplier_cancel_date,  # ship cancel date
                "MUNDRA",  # port of shipment
                dc_location,  # destination
                "USA",  # country
                dc_location,  # PORT OF LOADING
                mat_code,
                ordering_qty,  # Order quantity
                2250,  # plant
                item_number,  # customer material number
                pis,  # PIS
                "",  # PO avail date
                "KAVITA SUNDAKAR",  # A/C holder
                "", # Currency
                "2% commission to WUSA",  # notify
                "", # PO FILE NAME
                "", # PO FILE FORMAT
            ]
            self.macro_ws.append(macro_row)

        # apply stylings and create zip buffer.
        ft = Font(bold=True, size=12)
        format_header(
            self.ws,
            ft,
        )
        adjust_column_widths(self.ws)
        apply_borders_and_alignment(self.ws)
        format_number(self.ws, self.MASTER_SHEET_COLS["upc"])

        apply_borders_and_alignment(self.macro_ws)
        zip_buffer = BytesIO()

        with ZipFile(zip_buffer, "w") as zip_file:
            with BytesIO() as wbio:
                self.wb.save(wbio)
                file_name = f"WALMART_BEDSHEET_MASTER_SHEET_FILLED_{get_date()}.xlsx"
                zip_file.writestr(file_name, wbio.getvalue())

            with BytesIO() as macro_io:
                self.macro_wb.save(macro_io)
                zip_file.writestr(
                    "WALMART_MACRO_SHEET_BEDSHEET.xlsm", macro_io.getvalue()
                )

        zip_buffer.seek(0)

        return zip_buffer

    def _parse_po(self, ws: List[List]):
        """Parses the PO file and returns a dictionary with PO data"""
        # for each header (header row starts from row 2 (1)), get the colun name and set it's index
        for index, cell in enumerate(ws[1]):
            col_name = cell.lower().strip() if cell else None
            if col_name in self.PO_SHEET_COLS:
                self.PO_SHEET_COLS[col_name] = index

    def _parse_mastersheet(self, ws: Worksheet):
        """Parses the mastersheet to find relevant header columns and
        returns a dictionary of each roecord.
        """

        # for each header, get the colun name and set it's index
        for cell in ws[1]:
            col_name = cell.value.lower().strip() if cell.value else None

            if col_name in self.MASTER_SHEET_COLS:
                self.MASTER_SHEET_COLS[col_name] = cell.col_idx

        mastersheet_dict = {}

        for row in ws.iter_rows(min_row=2):
            if row[self.MASTER_SHEET_COLS["key"] - 1].value is None:
                break

            mastersheet_dict[row[self.MASTER_SHEET_COLS["key"] - 1].value] = [
                cell.value for cell in row
            ]

        return mastersheet_dict
