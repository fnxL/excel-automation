from app.common import Mastersheet
from typing import BinaryIO, List, Dict
from app.utils.helpers import (
    excel_cell_to_index,
    get_po_sheet,
    column_letter_to_idx,
    isEnd,
)
from app.utils.format import (
    adjust_column_widths,
    apply_borders_and_alignment,
    format_header,
    format_date,
    format_number,
    get_date,
)
from openpyxl import Workbook
from openpyxl.worksheet.worksheet import Worksheet
from openpyxl.styles import Font, PatternFill
from glob import glob
import os


class KohlsBedsheet(Mastersheet):
    MACRO_FILE_PATH = "macro/kohls-bedsheet.xlsm"
    MACRO_UPC_COL_IDX = 22
    MASTER_SHEET_COLS = {
        "upc": None,
        "qty": None,
        "po": None,
        "plant": None,
        "material number": None,
        "pis": None,
    }
    REPORT_HEADERS = [
        "PO",
        "PROGRAM",
        "SO",
        "TOTAL",
        "SW START",
        "SW END",
        "PACKING",
        "REMARKS",
        "STD PC 2PK",
        "KG PC 2PK",
        "KING SET",
        "QUEEN SET",
        "FULL SET",
        "CKING SET",
        "TWIN SET",
        "TWINXL SET",
        "FULL FIT",
        "TWINXL FIT",
        "TWIN FIT",
        "QUEEN FIT",
        "KING FIT",
        "C KING FIT",
    ]

    PO_DATA_CELLS = {
        "number": "C1",
        "sw_start": "I3",
        "sw_end": "I4",
        "packing": "N8",
        "upc_col": "M",
        "order_summary_header_row": 13,
        "qty_col": "F",
        "port_of_shipment": "J9",
    }

    def __init__(
        self,
        mastersheet: BinaryIO | str,
        po_folder: str = "uploads",
    ):
        super().__init__(
            mastersheet_cols=self.MASTER_SHEET_COLS,
            macro_file_path=self.MACRO_FILE_PATH,
            mastersheet=mastersheet,
        )
        self.po_folder = po_folder
        self.report_data = [self.REPORT_HEADERS]

    def process(self):
        for po_file in glob(os.path.join(self.po_folder, "*.xls*")):
            self._create_mastersheet(po_filepath=po_file)

        apply_borders_and_alignment(self.macro_ws)
        format_number(self.macro_ws, self.MACRO_UPC_COL_IDX)

        self._add_wb_to_zip(self.macro_wb, "KOHLS_MACRO_SHEET_BEDSHEET.xlsm")
        report_wb = self._create_report()
        self._add_wb_to_zip(report_wb, f"TOTAL_PO_QTY_BEDSHEET_{get_date()}.xlsx")
        self.report_data = [self.REPORT_HEADERS]  # Reset report data

    def _create_report(self):
        """Generates the Total PO QTY report"""
        wb = Workbook()
        ws = wb.active

        for row in self.report_data:
            ws.append(row)

        ft = Font(name="Calibri", bold=True, size=12)
        yellow_fill = PatternFill(
            start_color="FFFF00", end_color="FFFF00", fill_type="solid"
        )
        format_header(ws, ft, yellow_fill)
        apply_borders_and_alignment(ws)

        # Format columns "M" and "N"
        format_date(ws, 5, 6)
        adjust_column_widths(ws, 1.2)

        return wb

    def _create_mastersheet(self, po_filepath: str):
        po_sheet = get_po_sheet(po_filepath)
        po_data = self._parse_po(po_sheet)
        wb = Workbook()
        ws = wb.active
        ws.append(self.mastersheet_header)
        order_count = 0
        for row in po_sheet[po_data["po_header_row"] + 1 :]:
            if isEnd(row):
                break

            order_count += 1
            item_upc = int(row[column_letter_to_idx(self.PO_DATA_CELLS["upc_col"]) - 1])
            qty = int(row[column_letter_to_idx(self.PO_DATA_CELLS["qty_col"]) - 1])

            mastersheet_row = self.mastersheet_dict.get(
                item_upc, ["#N/A"] * len(self.mastersheet_header)
            )
            if mastersheet_row[self.MASTER_SHEET_COLS["upc"] - 1] == "#N/A":
                mastersheet_row[self.MASTER_SHEET_COLS["upc"] - 1] = item_upc
                mastersheet_row[self.MASTER_SHEET_COLS["upc"]] = (
                    "UPC not found in mastersheet"
                )

            mastersheet_row[self.MASTER_SHEET_COLS["qty"] - 1] = qty
            mastersheet_row[self.MASTER_SHEET_COLS["po"] - 1] = int(
                po_data["po_number"]
            )

            ws.append(mastersheet_row)

        ft = Font(bold=True, size=12)
        gray_fill = PatternFill(
            start_color="c0c0c0", end_color="c0c0c0", fill_type="solid"
        )
        format_header(ws, ft, gray_fill)
        apply_borders_and_alignment(ws)
        format_number(ws, self.MASTER_SHEET_COLS["upc"])
        adjust_column_widths(ws)

        self._create_report_row(po_sheet, po_data, order_count)
        self._add_to_macro(ws, po_data)
        self._add_wb_to_zip(
            wb, f"KOHLS_BEDSHEET_{po_data["po_number"]}_SO_SHEET_{get_date()}.xlsx"
        )

    def _add_to_macro(self, ws: Worksheet, po_data: dict):
        """Adds the data from newly created mastersheet
        into the macro sheet.
        """
        channel_type = "ECOM" if po_data["po_packing"] == "ECOM" else "RETAIL"
        sub_channel_type = "BRICK & MORTAR E" if channel_type == "ECOM" else "NIL"

        for row in ws.iter_rows(min_row=2):
            macro_row = [
                int(po_data["po_number"]),  # PO
                "",  # SO
                102083,  # SOLD TO PARTY
                102083,  # SHIP TO PARTY
                "W137",  # PAYMENT TERM
                "FOB",  # INCO TERMS
                po_data["po_port_of_shipment"],  # INCO TERM 2
                100023,  # end customer
                "YES",  # E2E CLARITY
                channel_type,
                sub_channel_type,
                "REPLENISHMENT",  # order type
                po_data["po_sw_start"],  # ship start date
                po_data["po_sw_end"],  # ship cancel date
                po_data["po_port_of_shipment"],  # port of shipment
                "SAVANNAH",  # destinatoin
                "USA",  # country
                "SAVANNAH",  # port of discharge
                row[self.MASTER_SHEET_COLS["material number"] - 1].value,  # mat code
                row[self.MASTER_SHEET_COLS["qty"] - 1].value,  # order qty
                row[self.MASTER_SHEET_COLS["plant"] - 1].value,  # plant
                row[self.MASTER_SHEET_COLS["upc"] - 1].value,  # customer material
                row[self.MASTER_SHEET_COLS["pis"] - 1].value,  # PIS
                "",  # PO AVAIL DATE
                "Akanksha Mehra",  # A/C HOLDER
            ]
            self.macro_ws.append(macro_row)

        self.macro_ws.append([""])

    def _create_report_row(self, po_sheet: List[List], po_data: Dict, order_count: int):
        """Creates a total qty report row with data from the PO file
        and appends it to the total qty sheet
        """
        report_row = [""] * len(self.REPORT_HEADERS)
        report_row[0] = po_data["po_number"]
        report_row[1] = po_data["po_program_name"]
        report_row[4] = po_data["po_sw_start"]
        report_row[5] = po_data["po_sw_end"]
        report_row[6] = po_data["po_packing"]
        order_summary = po_sheet[self.PO_DATA_CELLS["order_summary_header_row"] - 1]
        for i in range(3, len(order_summary)):
            cell_value = order_summary[i]
            offset = self.PO_DATA_CELLS["order_summary_header_row"] + order_count
            if cell_value == "":
                report_row[3] = po_sheet[offset][i]
                break

            cell_value_clean = cell_value.split("(")[0].strip()
            report_col_idx = self.REPORT_HEADERS.index(cell_value_clean.upper())
            report_row[report_col_idx] = po_sheet[offset][i]

        self.report_data.append(report_row)

    def _parse_po(self, ws: List[List]):
        """Parses the PO file and returns a dictionary with the PO data"""
        idxs = excel_cell_to_index(self.PO_DATA_CELLS["number"])
        po_number = ws[idxs[0]][idxs[1]]
        idxs = excel_cell_to_index(self.PO_DATA_CELLS["sw_start"])
        po_sw_start = ws[idxs[0]][idxs[1]]
        idxs = excel_cell_to_index(self.PO_DATA_CELLS["sw_end"])
        po_sw_end = ws[idxs[0]][idxs[1]]
        idxs = excel_cell_to_index(self.PO_DATA_CELLS["packing"])
        po_packing = ws[idxs[0]][idxs[1]]
        po_packing = "ECOM" if po_packing else "BULK"
        idxs = excel_cell_to_index(self.PO_DATA_CELLS["port_of_shipment"])
        port_of_shipment = ws[idxs[0]][idxs[1]]

        # find header row and program name
        po_header_row = None
        for index, row in enumerate(ws):
            if (
                row[column_letter_to_idx(self.PO_DATA_CELLS["upc_col"]) - 1].lower()
                == "item upc"
            ):
                po_header_row = index
                break

        return {
            "po_number": int(po_number),
            "po_sw_start": po_sw_start,
            "po_sw_end": po_sw_end,
            "po_packing": po_packing,
            "po_program_name": ws[po_header_row + 1][column_letter_to_idx("C") - 1],
            "po_header_row": po_header_row,
            "po_port_of_shipment": port_of_shipment,
        }
