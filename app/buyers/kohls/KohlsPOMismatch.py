from typing import BinaryIO
from glob import glob
from openpyxl import load_workbook
from openpyxl.worksheet.worksheet import Worksheet
from app.buyers.kohls.KohlsPDF import *
import os
from io import BytesIO
from zipfile import ZipFile


class KohlsPOMismatch:
    SHEET_COLS = {
        "po #": None,
        "customer material number": None,
        "po mismatch": None,
        "po qty": None,
        "po price": None,
        "order quantity": None,
        "net price": None,
    }

    def __init__(
        self,
        mastersheet: BinaryIO,
        filename: str,
        po_folder: str = "uploads",
    ):
        self.po_folder = po_folder
        self.filename = filename
        self.mastersheet_wb = load_workbook(filename=mastersheet)
        self.mastersheet_ws = self.mastersheet_wb.worksheets[0]
        # add three columns
        max_col = self.mastersheet_ws.max_column
        self.mastersheet_ws.cell(row=1, column=max_col + 1, value="PO Qty")
        self.mastersheet_ws.cell(row=1, column=max_col + 2, value="PO Price")
        self.mastersheet_ws.cell(row=1, column=max_col + 3, value="PO Mismatch")
        self.parse_mastersheet(self.mastersheet_ws)

    def parse_mastersheet(self, mastersheet: Worksheet):
        for cell in mastersheet[1]:
            col_name = cell.value.lower() if cell.value else None
            if col_name in self.SHEET_COLS:
                self.SHEET_COLS[col_name] = cell.col_idx
        return

    def process(self):
        current_po = None
        current_po_sheet = None
        for row in self.mastersheet_ws.iter_rows(min_row=2):
            po = row[self.SHEET_COLS["po #"] - 1].value
            if current_po != po:
                current_po = po
                current_po_sheet = self.read_po(po)

            upc = int(row[self.SHEET_COLS["customer material number"] - 1].value)
            qty = row[self.SHEET_COLS["order quantity"] - 1].value
            net_price = row[self.SHEET_COLS["net price"] - 1].value

            line_item = self.find_line_item(current_po_sheet, upc)
            row[self.SHEET_COLS["po qty"] - 1].value = line_item[3]
            row[self.SHEET_COLS["po price"] - 1].value = float(line_item[4])

            if line_item[3] != qty and line_item[4] != net_price:
                row[self.SHEET_COLS["po mismatch"] - 1].value = (
                    "Both QTY and Price mismatch"
                )
            elif line_item[3] != qty:
                row[self.SHEET_COLS["po mismatch"] - 1].value = "QTY mismatch"
            elif float(line_item[4]) != net_price:
                row[self.SHEET_COLS["po mismatch"] - 1].value = "Price mismatch"

        zip_buffer = BytesIO()
        with ZipFile(zip_buffer, "a") as zip_file:
            with BytesIO() as wb_io:
                self.mastersheet_wb.save(wb_io)
                zip_file.writestr(f"{self.filename}", wb_io.getvalue())
        zip_buffer.seek(0)
        return zip_buffer

    def read_po(self, po: int):
        po_file = f"{self.po_folder}/{po}.pdf"
        if not os.path.exists(po_file):
            raise FileNotFoundError(f"PO file {po_file} does not exist.")

        po_sheet = extract_table_rows(pdf_path=po_file)
        return po_sheet

    def find_line_item(self, po_sheet, upc):
        for row in po_sheet:
            if row[-1] == upc:
                return row
