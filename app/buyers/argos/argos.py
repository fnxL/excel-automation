import os
from io import BytesIO
from openpyxl import Workbook, load_workbook
from glob import glob
from openpyxl.worksheet.worksheet import Worksheet
from app.utils.helpers import (
    isEnd,
    get_po_sheet,
    get_date,
    adjust_column_widths,
    format_header,
    apply_styles,
    format_number,
    column_letter_to_idx,
    excel_cell_to_index,
    format_date,
)

from openpyxl.styles import Font, PatternFill, Border, Side


class Argos:

    SHEET_COLS = {
        "item number": None,
        "commodity": None,
        "ordered quantity": None,
        "buying price": None,
        "original origin delivery window start": None, 
    }

    HEADERS = [
        "PO#",
        "Item Number",
        "Commodity",
        "Ordered Quantity",
        "Buying Price",
        "Original Origin Delivery Window Start",
    ]

    def __init__(
        self,
    ):
        self.wb = Workbook()
        self.ws = self.wb.active
        self.ws.append(self.HEADERS)

    def process(self, po_folder: str = "uploads"):
        for po_file in glob(os.path.join(po_folder, "*.xls*")):
            po = os.path.basename(po_file).split(".")[0]
            wb = load_workbook(filename=po_file)
            ws = wb.active
            self._parse_po(ws)
            self.add_row(po, ws)
        # format and save it
        ft = Font(name="Calibri", bold=True, size=12)
        yellow_fill = PatternFill(
            start_color="FFFF00", end_color="FFFF00", fill_type="solid"
        )
        format_header(self.ws, ft, yellow_fill)
        apply_styles(self.ws)
        format_date(self.ws, 6, 6)
        adjust_column_widths(self.ws, 1.2)
        self.wb.save("argos.xlsx")

    def add_row(self, po, ws: Worksheet):
        for row in ws.iter_rows(min_row=2):
            item_number = row[self.SHEET_COLS["item number"] - 1].value
            commodity = row[self.SHEET_COLS["commodity"] - 1].value
            ordered_quantity = row[self.SHEET_COLS["ordered quantity"] - 1].value
            buying_price = row[self.SHEET_COLS["buying price"] - 1].value
            original_origin_delivery_window_start = row[
                self.SHEET_COLS["original origin delivery window start"] - 1
            ].value
            row_data = [
                int(po),
                item_number,
                commodity,
                ordered_quantity,
                buying_price,
                original_origin_delivery_window_start,
            ]
            self.ws.append(row_data)

    def _parse_po(self, ws: Worksheet):
        for cell in ws[1]:
            col_name = cell.value.lower() if cell.value else None
            if col_name in self.SHEET_COLS:
                self.SHEET_COLS[col_name] = cell.col_idx

        return
