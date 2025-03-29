from openpyxl.worksheet.worksheet import Worksheet
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment, NamedStyle
from datetime import date


def format_date(ws: Worksheet, startcol=0, endcol=1, date_format="DD-MM-YYYY"):
    for column in ws.iter_cols(min_col=startcol, max_col=endcol):  # Columns M and N
        for cell in column:
            cell.number_format = date_format


def format_number(ws: Worksheet, col: int):
    for row in ws.iter_cols(min_col=col, max_col=col):
        for cell in row:
            cell.number_format = "0"
    return


def format_header(ws: Worksheet, ft: Font = None, fill: PatternFill = None):
    for cell in ws[1]:
        if ft is not None:
            cell.font = ft
        if fill is not None:
            cell.fill = fill
    return


def get_date():
    return date.today().strftime("%d.%m.%Y")


def apply_borders_and_alignment(ws: Worksheet):
    """Applies all sided borders and center alignment to all cells"""
    center_aligned_text = Alignment(horizontal="center", vertical="center")
    thin_border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )
    for row in ws.iter_rows(min_row=1, max_row=ws.max_row):
        for cell in row:
            cell.border = thin_border
            cell.alignment = center_aligned_text

    return


def adjust_column_widths(ws: Worksheet, adjustment_factor=1.05):
    """
    Adjusts the widths of columns in an openpyxl worksheet to approximate Excel's auto-fit.
    An adjustment factor is used to account for variable character widths in different fonts.
    Args:
    - ws: The worksheet to adjust column widths for.
    - adjustment_factor: A multiplier applied to each column's calculated width to approximate font size and type.
    """
    for column in ws.columns:
        max_length = max(len(str(cell.value)) for cell in column)
        column_letter = column[0].column_letter
        ws.column_dimensions[column_letter].width = (max_length + 2) * adjustment_factor

    return
