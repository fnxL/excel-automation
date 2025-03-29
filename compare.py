import sys
from openpyxl import load_workbook


def compare(file1, file2):

    wb1 = load_workbook(file1, keep_vba=True)
    wb2 = load_workbook(file2, keep_vba=True)
    ws1 = wb1.active
    ws2 = wb2.active

    for row in range(1, ws1.max_row + 1):
        for col in range(1, ws1.max_column + 1):
            cell1 = ws1.cell(row=row, column=col)
            cell2 = ws2.cell(row=row, column=col)
            if cell1.value != cell2.value:
                print(
                    f"Cell {cell1.coordinate} has different values: {cell1.value} and {cell2.value}"
                )
                print(cell1.column_letter, cell1.row)
                print(cell2.column_letter, cell2.row)
                return


if __name__ == "__main__":
    if len(sys.argv) != 3:
        print("Usage: python compare.py <path to excel file> <path to excel file>")
        sys.exit(1)

    file1 = sys.argv[1]
    file2 = sys.argv[2]
    compare(file1, file2)
