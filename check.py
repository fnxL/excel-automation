from app.buyers.kohls.KohlsPDF import *
from glob import glob
from openpyxl import load_workbook
from openpyxl.worksheet.worksheet import Worksheet
import os

SHEET_COLS = {
    "po #": None,
    "customer material number": None,
    "po mismatch": None,
    "po qty": None,
    "po price": None,
    "order quantity": None,
    "net price": None,
}


def parse_mastersheet(mastersheet: Worksheet):
    for cell in mastersheet[1]:
        col_name = cell.value.lower() if cell.value else None
        if col_name in SHEET_COLS:
            SHEET_COLS[col_name] = cell.col_idx
    return


def main():
    excel = "uploads/excel.xlsx"
    wb = load_workbook(excel)
    ws = wb.active
    parse_mastersheet(ws)
    current_po = None
    current_po_sheet = None
    for row in ws.iter_rows(min_row=2):
        po = row[SHEET_COLS["po #"] - 1].value
        if current_po != po:
            current_po = po
            current_po_sheet = read_po(po)

        upc = int(row[SHEET_COLS["customer material number"] - 1].value)
        qty = row[SHEET_COLS["order quantity"] - 1].value
        net_price = row[SHEET_COLS["net price"] - 1].value

        line_item = find_line_item(current_po_sheet, upc)
        row[SHEET_COLS["po qty"] - 1].value = line_item[3]
        row[SHEET_COLS["po price"] - 1].value = float(line_item[4])

        if line_item[3] != qty and line_item[4] != net_price:
            row[SHEET_COLS["po mismatch"] - 1].value = "Both QTY and Price mismatch"
        elif line_item[3] != qty:
            row[SHEET_COLS["po mismatch"] - 1].value = "QTY mismatch"
        elif float(line_item[4]) != net_price:
            row[SHEET_COLS["po mismatch"] - 1].value = "Price mismatch"

    wb.save("uploads/excel_updated.xlsx")
    wb.close()


def find_line_item(po_sheet, upc):
    for row in po_sheet:
        if row[-1] == upc:
            return row


def read_po(po: int):
    po_file = f"uploads/{po}.pdf"
    if not os.path.exists(po_file):
        raise FileNotFoundError(f"PO file {po_file} does not exist.")

    po_sheet = extract_table_rows(pdf_path=po_file)
    return po_sheet


if __name__ == "__main__":
    main()
